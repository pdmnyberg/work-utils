import sys
import csv
import openpyxl
import argparse


def main(source: str, target_template: str):
    workbook = openpyxl.load_workbook(
        source,
        read_only=True,
        data_only=True
    )
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        headers = headers_from_sheet(sheet)
        data = list(sheet_to_dict(headers, sheet))
        with open(target_template.format(sheet=sheet_name), "w") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)


def headers_from_sheet(sheet):
    return [h.value.strip() for h in next(sheet.iter_rows())]


def sheet_to_dict(headers: list[str], sheet):
    for row in sheet.iter_rows(min_row=2):
        cell_values = [
            None
            if cell.value is None
            else str(cell.value).strip()
            for cell in row
        ]
        yield {
            key: value
            for key, value in zip(headers, cell_values)
        }


main(*sys.argv[1:])